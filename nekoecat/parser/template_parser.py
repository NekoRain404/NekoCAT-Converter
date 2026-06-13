"""
Template xlsx parser — reads an SSC Tool .xlsx template into structured data.

This parser reads the SSC Application Description xlsx format so we can
use it as a reference template for generating output.

Only depends on: openpyxl.  No model/ imports (returns plain dicts).
"""
from pathlib import Path
from typing import Any
from rich.console import Console

console = Console()


def parse_template_xlsx(filepath: str) -> dict[str, Any]:
    """Parse an SSC Tool template .xlsx into a dict of sheet data.

    Returns:
        {
            "device_info": {...},          # from "Device Info" or first sheet
            "object_list": [{...}, ...],   # from "Object List" sheet
            "subindex_list": [{...}, ...], # from "Subindex List" sheet
            "sheet_names": [...],          # all sheet names
        }
    """
    import openpyxl

    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Template xlsx not found: {path}")

    wb = openpyxl.load_workbook(str(path), data_only=True)
    result: dict[str, Any] = {"sheet_names": wb.sheetnames}

    # ── Device Info ────────────────────────────
    result["device_info"] = _parse_device_info(wb)

    # ── Object List ────────────────────────────
    result["object_list"] = _parse_table_sheet(wb, ["Object List", "Objects", "ObjList"])

    # ── Subindex List ──────────────────────────
    result["subindex_list"] = _parse_table_sheet(wb, ["Subindex List", "SubIndex", "SubItems"])

    wb.close()

    console.print(
        f"[green]Template loaded:[/green] "
        f"{len(result['object_list'])} objects, "
        f"{len(result['subindex_list'])} subindices"
    )
    return result


def _parse_device_info(wb) -> dict[str, str]:
    """Read device info from the 'Device Info' or first sheet."""
    info: dict[str, str] = {}

    sheet = None
    for name in ["Device Info", "DeviceInfo", "Device", "Info"]:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None and wb.sheetnames:
        sheet = wb[wb.sheetnames[0]]

    if sheet is None:
        return info

    for row in sheet.iter_rows(min_row=1, max_row=50, values_only=False):
        if len(row) >= 2 and row[0].value is not None:
            key = str(row[0].value).strip()
            val = str(row[1].value or "").strip()
            if key:
                info[key] = val

    return info


def _parse_table_sheet(wb, candidate_names: list[str]) -> list[dict[str, str]]:
    """Read a table-style sheet (header row + data rows) into a list of dicts."""
    sheet = None
    for name in candidate_names:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        return []

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    # First row = header
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]

    result: list[dict[str, str]] = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        entry = {}
        for i, val in enumerate(row):
            if i < len(headers):
                entry[headers[i]] = str(val).strip() if val is not None else ""
        result.append(entry)

    return result
