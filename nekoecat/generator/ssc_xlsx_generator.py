"""
SSC Tool xlsx generator — writes a DeviceModel into the SSC Application
Description xlsx format that SSC Tool can import.

Column layout follows dev doc §11:
  Index, ObjectCode, SI, DataType, Name, Default Value,
  Min Value, Max Value, M/O/C, B/S, Access, rx/tx,
  CoeRead, CoeWrite, Description

Critical rule: CoeRead/CoeWrite are only on Object rows.
               Entry rows must have CoeRead/CoeWrite = empty.

Only depends on: model/, constants.  No parser/engine imports.
"""
from pathlib import Path
from datetime import datetime
from typing import Optional
from rich.console import Console
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from nekoecat.model import DeviceModel, ObjectEntry
from nekoecat.constants import COMM_OBJECT_RANGE

console = Console()

# ── style ─────────────────────────────────────
_HEADER_FONT = Font(bold=True, size=10)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

# Dev doc §11 column order
SSC_COLUMNS = [
    "Index", "ObjectCode", "SI", "DataType", "Name",
    "Default Value", "Min Value", "Max Value",
    "M/O/C", "B/S", "Access", "rx/tx",
    "CoeRead", "CoeWrite", "Description",
]


def generate_ssc_xlsx(
    device: DeviceModel,
    output_path: str,
    template_path: Optional[str] = None,
) -> str:
    """Generate an SSC Tool-compatible xlsx file.

    Args:
        device: The fully processed DeviceModel.
        output_path: Where to write the xlsx file.
        template_path: Optional template to copy structure from.

    Returns:
        Absolute path of the generated file.
    """
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    if template_path and Path(template_path).exists():
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()

    _write_device_info(wb, device)
    _write_objects_sheet(wb, device)
    _write_subindices_sheet(wb, device)

    wb.save(str(output))
    wb.close()
    console.print(f"[green]Generated:[/green] {output}")
    return str(output.resolve())


# ──────────────────────────────────────────────
# Sheet: Device Info
# ──────────────────────────────────────────────

def _write_device_info(wb, device: DeviceModel):
    ws = _ensure_sheet(wb, "Device Info")
    ws.delete_rows(1, ws.max_row)

    ident = device.identity
    rows = [
        ("Vendor ID", f"0x{ident.vendor_id:08X}"),
        ("Vendor Name", ident.vendor_name),
        ("Product Code", f"0x{ident.product_code:08X}"),
        ("Revision Number", f"0x{ident.revision_number:08X}"),
        ("Device Name", ident.device_name),
        ("Device Type", str(ident.device_type)),
        ("Hardware Version", ident.hardware_version),
        ("Software Version", ident.software_version),
        ("Generated At", datetime.now().isoformat()),
        ("Generator", "NekoECAT Converter v0.1.0"),
    ]
    for i, (key, val) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=key).font = _HEADER_FONT
        ws.cell(row=i, column=2, value=val)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40


# ──────────────────────────────────────────────
# Sheet: Objects (one row per object, with CoE flags)
# ──────────────────────────────────────────────

def _write_objects_sheet(wb, device: DeviceModel):
    """Write object-level rows. CoeRead/CoeWrite are set here ONLY."""
    ws = _ensure_sheet(wb, "Object List")
    ws.delete_rows(1, ws.max_row)
    _write_header(ws, SSC_COLUMNS)

    row = 2
    for obj in device.sorted_objects():
        if obj.index in COMM_OBJECT_RANGE:
            continue
        if obj.skipped:
            continue

        _write_row(ws, row, _object_to_row(obj))
        row += 1

    _auto_width(ws)


# ──────────────────────────────────────────────
# Sheet: Subindices (one row per subindex, NO CoE flags)
# ──────────────────────────────────────────────

def _write_subindices_sheet(wb, device: DeviceModel):
    """Write subindex-level rows. CoeRead/CoeWrite are EMPTY here."""
    ws = _ensure_sheet(wb, "Subindex List")
    ws.delete_rows(1, ws.max_row)
    _write_header(ws, SSC_COLUMNS)

    row = 2
    for obj in device.sorted_objects():
        if obj.index in COMM_OBJECT_RANGE:
            continue
        if not obj.subindices:
            continue
        if obj.skipped:
            continue

        for sub in obj.subindices:
            if sub.skipped:
                continue
            _write_row(ws, row, _subindex_to_row(obj, sub))
            row += 1

    _auto_width(ws)


# ──────────────────────────────────────────────
# Row builders (dev doc §11)
# ──────────────────────────────────────────────

def _object_to_row(obj: ObjectEntry) -> dict[str, str]:
    """Build an object-level row dict.  CoE flags go here."""
    return {
        "Index": f"0x{obj.index:04X}",
        "ObjectCode": _object_code_str(obj.object_type),
        "SI": "",
        "DataType": "",
        "Name": obj.name,
        "Default Value": "",
        "Min Value": "",
        "Max Value": "",
        "M/O/C": "O",
        "B/S": "S",
        "Access": "",
        "rx/tx": "",
        "CoeRead": "TRUE" if obj.coe_read is True else "",
        "CoeWrite": "TRUE" if obj.coe_write is True else "",
        "Description": "; ".join([obj.source]) if obj.source else "",
    }


def _subindex_to_row(obj: ObjectEntry, sub) -> dict[str, str]:
    """Build a subindex-level row.  CoE flags are EMPTY here."""
    return {
        "Index": f"0x{obj.index:04X}",
        "ObjectCode": "",
        "SI": sub.subindex,
        "DataType": sub.data_type or obj.data_type,
        "Name": sub.name,
        "Default Value": sub.default_value or obj.default_value,
        "Min Value": "",
        "Max Value": "",
        "M/O/C": "O",
        "B/S": "S",
        "Access": _to_ssc_access(sub.access or obj.access),
        "rx/tx": sub.pdo_mapping or "",
        "CoeRead": "",   # MUST be empty at entry level
        "CoeWrite": "",  # MUST be empty at entry level
        "Description": sub.original_data_type or "",
    }


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def _object_code_str(code: int) -> str:
    return {7: "VAR", 8: "ARRAY", 9: "RECORD"}.get(code, str(code))


def _to_ssc_access(raw: str) -> str:
    """Normalize access strings to SSC Tool format."""
    s = raw.strip().lower()
    if s in ("ro", "r"):
        return "ro"
    if s in ("rw", "r/w"):
        return "rw"
    if s in ("wo", "w"):
        return "wo"
    if s == "const":
        return "ro"
    return raw


def _ensure_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def _write_header(ws, columns: list[str]):
    for col, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _write_row(ws, row_num: int, data: dict[str, str]):
    for col_idx, col_name in enumerate(SSC_COLUMNS, start=1):
        ws.cell(row=row_num, column=col_idx, value=str(data.get(col_name, "")))


def _auto_width(ws):
    for col in range(1, len(SSC_COLUMNS) + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            val = row[0]
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 3, 30)
